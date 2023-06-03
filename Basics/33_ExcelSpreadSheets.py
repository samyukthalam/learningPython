import openpyxl as xl
from openpyxl.chart import BarChart, Reference

file_name="C:\\Users\\admin\\PycharmProjects\\LearnPython1\\Files\\ExcelData.xlsx"


def process_workbook(file_name):
        wb = xl.load_workbook(file_name)
        sheet = wb['Sheet1']

        for row in range(2,sheet.max_row+1):
            cell=sheet.cell(row,3)
            corrected_price=cell.value*0.9
            corrected_price_cell=sheet.cell(row,4)
            corrected_price_cell.value=corrected_price

        chart_Values=Reference(sheet,
                              min_row=2,
                              max_row=sheet.max_row,
                              min_col=4,
                              max_col=4)
        chart=BarChart()
        chart.add_data(chart_Values)
        sheet.add_chart(chart,'e2')

        wb.save('C:\\Users\\admin\\PycharmProjects\\LearnPython1\\Files\\ExcelData_New.xlsx')

process_workbook(file_name)